import openpyxl


# Retrieve the row number
def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


# Retrieve the column number
def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


# Retrieve any specific cell data
def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


# Write on cell
def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


# Retrieve all data from Excel
def get_data_from_excel(path, sheet_name):

    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list
